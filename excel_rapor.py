import os
import calendar
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from tkinter import messagebox

def excel_dolu_puantaj(secilen_yil, secilen_ay, personeller, veri_getir_func):
    """Masaüstüne veritabanı kayıtlarına göre kurumsal logolu, görev sütunlu ve mesai renkli puantaj üretir."""
    excel_rapor_ust_seviye(secilen_yil, secilen_ay, personeller, veri_getir_func, rapor_tipi="dolu")

def excel_bos_puantaj(secilen_yil, secilen_ay, personeller):
    """Masaüstüne içi boş, şablon şeklinde kurumsal logolu ve görev sütunlu puantaj üretir."""
    excel_rapor_ust_seviye(secilen_yil, secilen_ay, personeller, veri_getir_func=None, rapor_tipi="bos")

def excel_rapor_ust_seviye(secilen_yil, secilen_ay, personeller, veri_getir_func=None, rapor_tipi="dolu"):
    try:
        # 1. Masaüstü Yolunu Güvenli Şekilde Bulma
        home = os.path.expanduser("~")
        desktop_path = os.path.join(home, "Desktop")
        if not os.path.exists(desktop_path):
            alt_path = os.path.join(home, "Masaüstü")
            if os.path.exists(alt_path):
                desktop_path = alt_path
            else:
                alt_path = os.path.join(home, "OneDrive", "Desktop")
                if os.path.exists(alt_path):
                    desktop_path = alt_path
                else:
                    alt_path = os.path.join(home, "OneDrive", "Masaüstü")
                    if os.path.exists(alt_path):
                        desktop_path = alt_path

        dosya_adi = f"Puantaj_Raporu_{secilen_yil}_{secilen_ay:02d}_{rapor_tipi}.xlsx"
        kayit_yolu = os.path.join(desktop_path, dosya_adi)

        wb = Workbook()
        ws = wb.active
        ws.title = f"Puantaj {secilen_ay:02d}-{secilen_yil}"
        ws.views.sheetView[0].showGridLines = True

        # 2. Kurumsal Renk ve Stil Tanımlamaları
        font_ana_baslik = Font(name="Segoe UI", size=14, bold=True, color="1F4E78")
        font_baslik = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        font_veri = Font(name="Segoe UI", size=10)
        font_bold = Font(name="Segoe UI", size=10, bold=True)
        font_imza = Font(name="Segoe UI", size=11, bold=True, color="333333")
        
        fill_baslik = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Kurumsal Lacivert
        fill_cumartesi = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft Sarı
        fill_pazar = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")     # Soft Kırmızı

        # Yoklama Durum Renkleri (Arayüzle Birebir Eşit)
        fill_durum_g = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Yeşil (Geldi)
        fill_durum_p = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid") # Turuncu (Pazar Geldi)
        fill_durum_y = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Kırmızı (Yok)
        fill_durum_i = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid") # Mavi (İzinli)
        fill_durum_r = PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid") # Koyu Sarı (Raporlu)

        # MESAİ YOĞUNLUK RENKLERİ (Isı Haritası - Skala)
        fill_mesai_az = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")   # 0 - 2 Saat Arası (Açık Sarı/Turuncu)
        fill_mesai_orta = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid") # 2 - 4 Saat Arası (Orta)
        fill_mesai_yogun = PatternFill(start_color="FCD34D", end_color="FCD34D", fill_type="solid") # 4 Saat Üzeri (Yoğun)

        # Yazı Renkleri
        font_durum_g = Font(name="Segoe UI", size=10, bold=True, color="22C55E")
        font_durum_p = Font(name="Segoe UI", size=10, bold=True, color="F59E0B")
        font_durum_y = Font(name="Segoe UI", size=10, bold=True, color="EF4444")
        font_durum_i = Font(name="Segoe UI", size=10, bold=True, color="3B82F6")
        font_durum_r = Font(name="Segoe UI", size=10, bold=True, color="EA580C")
        font_durum_normal = Font(name="Segoe UI", size=10, color="9CA3AF")

        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center")

        gun_sayisi = calendar.monthrange(secilen_yil, secilen_ay)[1]
        aylar_tr = {
            1: "OCAK", 2: "ŞUBAT", 3: "MART", 4: "NİSAN", 5: "MAYIS", 6: "HAZİRAN",
            7: "TEMMUZ", 8: "AĞUSTOS", 9: "EYLÜL", 10: "EKİM", 11: "KASIM", 12: "ARALIK"
        }
        secili_ay_adi = aylar_tr.get(secilen_ay, "")

        # 3. LOGO VE BAŞLIK ALANI
        ws.row_dimensions[2].height = 45
        
        if os.path.exists("logo-kibritci.png"):
            img = Image("logo-kibritci.png")
            img.width = 140
            img.height = 50
            ws.add_image(img, "A2")
        
        # Başlığı görev sütununu da kapsayacak şekilde 4. sütundan başlatıp birleştiriyoruz
        ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=gun_sayisi + 2)
        baslik_hucre = ws.cell(row=2, column=4, value=f"KİBRİTÇİ İNŞAAT - {secili_ay_adi} {secilen_yil} PUANTAJ EVRAKI")
        baslik_hucre.font = font_ana_baslik
        baslik_hucre.alignment = Alignment(horizontal="left", vertical="center")

        # 4. TABLO BAŞLIK SATIRI (SATIR 4)
        tablo_baslik_satir = 4
        ws.row_dimensions[tablo_baslik_satir].height = 28

        # Personel Adı Sütunu (A Sütunu)
        ws.cell(row=tablo_baslik_satir, column=1, value="Personel Adı Soyadı").font = font_baslik
        ws.cell(row=tablo_baslik_satir, column=1).fill = fill_baslik
        ws.cell(row=tablo_baslik_satir, column=1).alignment = align_left
        ws.cell(row=tablo_baslik_satir, column=1).border = thin_border
        ws.column_dimensions['A'].width = 25

        # GÖREV / ÜNVAN SÜTUNU (B Sütunu)
        ws.cell(row=tablo_baslik_satir, column=2, value="Görev / Ünvan").font = font_baslik
        ws.cell(row=tablo_baslik_satir, column=2).fill = fill_baslik
        ws.cell(row=tablo_baslik_satir, column=2).alignment = align_left
        ws.cell(row=tablo_baslik_satir, column=2).border = thin_border
        ws.column_dimensions['B'].width = 16

        gunler_kisa = ["Pzt", "Sal", "Çar", "Per", "Cum", "Cmt", "Paz"]

        # Günleri Başlığa Yazma (Görev eklendiği için C sütunundan başlıyor: g + 2)
        for g in range(1, gun_sayisi + 1):
            tarih = date(secilen_yil, secilen_ay, g)
            hafta_gunu = tarih.weekday()
            
            hucre = ws.cell(row=tablo_baslik_satir, column=g + 2, value=f"{g:02d}\n{gunler_kisa[hafta_gunu]}")
            hucre.font = font_baslik
            hucre.fill = fill_baslik
            hucre.alignment = align_center
            hucre.border = thin_border
            ws.column_dimensions[hucre.column_letter].width = 4.5

        col_toplam_gun = gun_sayisi + 3
        col_toplam_mesai = gun_sayisi + 4

        cell_tg = ws.cell(row=tablo_baslik_satir, column=col_toplam_gun, value="Toplam\nGün")
        cell_tg.font = font_baslik
        cell_tg.fill = fill_baslik
        cell_tg.alignment = align_center
        cell_tg.border = thin_border
        ws.column_dimensions[cell_tg.column_letter].width = 10

        cell_tm = ws.cell(row=tablo_baslik_satir, column=col_toplam_mesai, value="Toplam\nMesai")
        cell_tm.font = font_baslik
        cell_tm.fill = fill_baslik
        cell_tm.alignment = align_center
        cell_tm.border = thin_border
        ws.column_dimensions[cell_tm.column_letter].width = 10

        current_row = 5

        # 5. PERSONEL VERİ DÖNGÜSÜ
        for p in personeller:
            p_id = p.get("id")
            p_gorev = p.get("gorev", "İşçi") # Veritabanından gelen görev tanımı, yoksa varsayılan İşçi

            if "ad_soyad" in p:
                p_ad = p["ad_soyad"]
            else:
                p_ad = f"{p.get('ad', '')} {p.get('soyad', '')}".strip() or f"Personel (ID: {p_id})"

            gun_map = {}
            mesai_map = {}
            
            if rapor_tipi == "dolu" and veri_getir_func:
                veri = veri_getir_func(p_id, secilen_yil, secilen_ay)
                for v in veri:
                    try:
                        tarih_str = v.get("tarih")
                        gun_no = int(tarih_str.split("-")[2])
                        gun_map[gun_no] = v.get("durum")
                        mesai_map[gun_no] = float(v.get("mesai_saati") or 0)
                    except Exception:
                        continue

            # --- SATIR 1: YOKLAMA SATIRI ---
            ws.row_dimensions[current_row].height = 20
            
            # Personel Adı
            cell_name1 = ws.cell(row=current_row, column=1, value=p_ad)
            cell_name1.font = font_bold
            cell_name1.alignment = align_left
            cell_name1.border = thin_border

            # Personel Görevi
            cell_gorev1 = ws.cell(row=current_row, column=2, value=p_gorev)
            cell_gorev1.font = font_veri
            cell_gorev1.alignment = align_left
            cell_gorev1.border = thin_border

            toplam_gun = 0
            toplam_mesai = 0

            for g in range(1, gun_sayisi + 1):
                hucre_yoklama = ws.cell(row=current_row, column=g + 2)
                hucre_yoklama.alignment = align_center
                hucre_yoklama.border = thin_border
                
                hafta_gunu = date(secilen_yil, secilen_ay, g).weekday()
                durum = gun_map.get(g)

                if rapor_tipi == "dolu":
                    # Pazar Günü Geldiyse -> "P" ve Turuncu Arka Plan
                    if durum == "Geldi" and hafta_gunu == 6:
                        hucre_yoklama.value = "P"
                        hucre_yoklama.font = font_durum_p
                        hucre_yoklama.fill = fill_durum_p
                        toplam_gun += 1
                    # Normal Geldiyse -> "G" ve Yeşil Arka Plan
                    elif durum == "Geldi":
                        hucre_yoklama.value = "G"
                        hucre_yoklama.font = font_durum_g
                        hucre_yoklama.fill = fill_durum_g
                        toplam_gun += 1
                    # Yok -> "Y" ve Kırmızı Arka Plan
                    elif durum == "Yok":
                        hucre_yoklama.value = "Y"
                        hucre_yoklama.font = font_durum_y
                        hucre_yoklama.fill = fill_durum_y
                    # İzinli -> "İ" ve Mavi Arka Plan
                    elif durum == "İzinli":
                        hucre_yoklama.value = "İ"
                        hucre_yoklama.font = font_durum_i
                        hucre_yoklama.fill = fill_durum_i
                    # Raporlu -> "R" ve Sarı Arka Plan
                    elif durum == "Raporlu":
                        hucre_yoklama.value = "R"
                        hucre_yoklama.font = font_durum_r
                        hucre_yoklama.fill = fill_durum_r
                    else:
                        hucre_yoklama.value = "-"
                        hucre_yoklama.font = font_durum_normal
                        if hafta_gunu == 5: hucre_yoklama.fill = fill_cumartesi
                        elif hafta_gunu == 6: hucre_yoklama.fill = fill_pazar
                else:
                    hucre_yoklama.value = ""
                    if hafta_gunu == 5: hucre_yoklama.fill = fill_cumartesi
                    elif hafta_gunu == 6: hucre_yoklama.fill = fill_pazar

            cell_tg_val = ws.cell(row=current_row, column=col_toplam_gun, value=f"{toplam_gun} Gün" if rapor_tipi == "dolu" else "")
            cell_tg_val.font = font_bold
            cell_tg_val.alignment = align_center
            cell_tg_val.border = thin_border
            ws.cell(row=current_row, column=col_toplam_mesai).border = thin_border

            current_row += 1

            # --- SATIR 2: MESAİ SATIRI ---
            ws.row_dimensions[current_row].height = 18
            
            cell_name2 = ws.cell(row=current_row, column=1, value=" ⏱ Mesai Saati")
            cell_name2.font = Font(name="Segoe UI", size=9, italic=True, color="595959")
            cell_name2.alignment = align_left
            cell_name2.border = thin_border

            # Alt satırdaki görev sütununu boş bırakıp çizgisini çekiyoruz
            cell_gorev2 = ws.cell(row=current_row, column=2, value="")
            cell_gorev2.border = thin_border

            for g in range(1, gun_sayisi + 1):
                hucre_mesai = ws.cell(row=current_row, column=g + 2)
                hucre_mesai.alignment = align_center
                hucre_mesai.border = thin_border
                
                hafta_gunu = date(secilen_yil, secilen_ay, g).weekday()
                mesai = mesai_map.get(g, 0.0)

                if rapor_tipi == "dolu" and mesai > 0:
                    hucre_mesai.value = mesai
                    hucre_mesai.font = font_bold
                    toplam_mesai += mesai
                    
                    # MESAİ YOĞUNLUK SKALASI (Isı Haritası) ENTEGRASYONU
                    if mesai <= 2.0:
                        hucre_mesai.fill = fill_mesai_az
                    elif mesai <= 4.0:
                        hucre_mesai.fill = fill_mesai_orta
                    else:
                        hucre_mesai.fill = fill_mesai_yogun
                else:
                    hucre_mesai.value = ""
                    if hafta_gunu == 5: hucre_mesai.fill = fill_cumartesi
                    elif hafta_gunu == 6: hucre_mesai.fill = fill_pazar

            ws.cell(row=current_row, column=col_toplam_gun).border = thin_border
            cell_tm_val = ws.cell(row=current_row, column=col_toplam_mesai, value=f"{round(toplam_mesai, 1)} Saat" if rapor_tipi == "dolu" else "")
            cell_tm_val.font = font_bold
            cell_tm_val.alignment = align_center
            cell_tm_val.border = thin_border

            current_row += 1

        # 6. KURUMSAL İMZA BARS ALANI
        current_row += 3
        
        imza_unvanlar = ["HAZIRLAYAN", "MUHASEBE", "ŞANTİYE ŞEFİ", "PROJE MÜDÜRÜ"]
        sutun_araligi = max(1, gun_sayisi // 4)
        
        for i, unvan in enumerate(imza_unvanlar):
            start_col = 1 + (i * sutun_araligi)
            end_col = start_col + sutun_araligi - 1
            if end_col > col_toplam_mesai:
                end_col = col_toplam_mesai
                
            ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=end_col)
            imza_cell = ws.cell(row=current_row, column=start_col, value=unvan)
            imza_cell.font = font_imza
            imza_cell.alignment = align_center
            
            ws.merge_cells(start_row=current_row + 3, start_column=start_col, end_row=current_row + 3, end_column=end_col)
            cizgi_cell = ws.cell(row=current_row + 3, column=start_col, value="...................................")
            cizgi_cell.font = font_veri
            cizgi_cell.alignment = align_center

        # 7. Kaydet ve Bitir
        wb.save(kayit_yolu)
        messagebox.showinfo("Başarılı", f"Kurumsal Excel raporu başarıyla masaüstüne kaydedildi:\n{dosya_adi}")

    except Exception as e:
        messagebox.showerror("Hata", f"Excel raporu oluşturulurken bir hata meydana geldi:\n{str(e)}")
def excel_bos_puantaj(secilen_yil, secilen_ay, personeller):
    """
    Şantiyede elle doldurulmak üzere müsvette/şablon formatında,
    A3 boyutunda, Yatay, her sayfada tarih cetvelini tekrarlayan (Print Titles),
    logolu, görev sütunlu ve boş hücreli kurumsal puantaj üretir.
    """
    try:
        import os
        import calendar
        from datetime import date
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.drawing.image import Image
        from tkinter import messagebox

        # 1. Masaüstü Yolunu Güvenli Şekilde Bulma
        home = os.path.expanduser("~")
        desktop_path = os.path.join(home, "Desktop")
        if not os.path.exists(desktop_path):
            alt_path = os.path.join(home, "Masaüstü")
            if os.path.exists(alt_path):
                desktop_path = alt_path
            else:
                alt_path = os.path.join(home, "OneDrive", "Desktop")
                if os.path.exists(alt_path):
                    desktop_path = alt_path
                else:
                    alt_path = os.path.join(home, "OneDrive", "Masaüstü")
                    if os.path.exists(alt_path):
                        desktop_path = alt_path

        dosya_adi = f"Bos_Puantaj_Sablonu_{secilen_yil}_{secilen_ay:02d}.xlsx"
        kayit_yolu = os.path.join(desktop_path, dosya_adi)

        wb = Workbook()
        ws = wb.active
        ws.title = f"Boş Şablon {secilen_ay:02d}-{secilen_yil}"
        ws.views.sheetView[0].showGridLines = True

        # =========================================================================
        # --- A3 FORMATI, YATAY BASKI VE SAYFA TEKRARLAMA AYARLARI (GÜNCEL) ---
        # =========================================================================
        ws.page_setup.paperSize = 8  # 8 değeri doğrudan evrensel A3 kağıt boyutunu temsil eder
        ws.page_setup.orientation = 'landscape'  # Doğrudan küçük harflerle metin olarak verdik, Türkçe İ harfine dönüşemez
        
        # Sütunların sayfaya tam sığması için ölçeklendirme
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1   # Tüm tarih sütunlarını yatayda TEK sayfaya sığdır
        ws.page_setup.fitToHeight = 0  # Dikeyde personel sayısına göre aşağı doğru sayfa sayfa uzayabilir
        
        # MÜKEMMEL ÇIKTI İÇİN CRITICAL AYAR: Sayfa ayrıldığında üstteki ilk 4 satırı (Logo ve Tarih Cetveli) her sayfaya basar!
        ws.print_title_rows = '1:4'
        
        # Kenar boşlukları (A3 kağıdı maksimum verimle kullanmak için daraltıldı)
        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        # =========================================================================

        # 2. Kurumsal Stil ve Renk Tanımlamaları
        font_ana_baslik = Font(name="Segoe UI", size=14, bold=True, color="1F4E78")
        font_baslik = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        font_veri = Font(name="Segoe UI", size=10)
        font_bold = Font(name="Segoe UI", size=10, bold=True)
        font_imza = Font(name="Segoe UI", size=11, bold=True, color="333333")
        
        fill_baslik = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Kurumsal Lacivert
        fill_cumartesi = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft Sarı
        fill_pazar = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")     # Soft Kırmızı

        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center")

        gun_sayisi = calendar.monthrange(secilen_yil, secilen_ay)[1]
        aylar_tr = {
            1: "OCAK", 2: "ŞUBAT", 3: "MART", 4: "NİSAN", 5: "MAYIS", 6: "HAZİRAN",
            7: "TEMMUZ", 8: "AĞUSTOS", 9: "EYLÜL", 10: "EKİM", 11: "KASIM", 12: "ARALIK"
        }
        secili_ay_adi = aylar_tr.get(secilen_ay, "")

        # 3. LOGO VE BAŞLIK ALANI (Satır 1-3 arası)
        ws.row_dimensions[2].height = 45
        
        if os.path.exists("logo-kibritci.png"):
            img = Image("logo-kibritci.png")
            img.width = 140
            img.height = 50
            ws.add_image(img, "A2")
        
        ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=gun_sayisi + 2)
        baslik_hucre = ws.cell(row=2, column=4, value=f"KİBRİTÇİ İNŞAAT - {secili_ay_adi} {secilen_yil} BOŞ PUANTAJ ŞABLONU")
        baslik_hucre.font = font_ana_baslik
        baslik_hucre.alignment = Alignment(horizontal="left", vertical="center")

        # 4. TABLO BAŞLIK SATIRI (SATIR 4 - TARİH CETVELİ)
        tablo_baslik_satir = 4
        ws.row_dimensions[tablo_baslik_satir].height = 28

        # Personel Adı
        ws.cell(row=tablo_baslik_satir, column=1, value="Personel Adı Soyadı").font = font_baslik
        ws.cell(row=tablo_baslik_satir, column=1).fill = fill_baslik
        ws.cell(row=tablo_baslik_satir, column=1).alignment = align_left
        ws.cell(row=tablo_baslik_satir, column=1).border = thin_border
        ws.column_dimensions['A'].width = 25

        # Görev Sütunu
        ws.cell(row=tablo_baslik_satir, column=2, value="Görev / Ünvan").font = font_baslik
        ws.cell(row=tablo_baslik_satir, column=2).fill = fill_baslik
        ws.cell(row=tablo_baslik_satir, column=2).alignment = align_left
        ws.cell(row=tablo_baslik_satir, column=2).border = thin_border
        ws.column_dimensions['B'].width = 16

        gunler_kisa = ["Pzt", "Sal", "Çar", "Per", "Cum", "Cmt", "Paz"]

        # Gün Başlıkları
        for g in range(1, gun_sayisi + 1):
            tarih = date(secilen_yil, secilen_ay, g)
            hafta_gunu = tarih.weekday()
            
            hucre = ws.cell(row=tablo_baslik_satir, column=g + 2, value=f"{g:02d}\n{gunler_kisa[hafta_gunu]}")
            hucre.font = font_baslik
            hucre.fill = fill_baslik
            hucre.alignment = align_center
            hucre.border = thin_border
            ws.column_dimensions[hucre.column_letter].width = 5.0 

        col_toplam_gun = gun_sayisi + 3
        col_toplam_mesai = gun_sayisi + 4

        # Toplam Sütun Başlıkları
        cell_tg = ws.cell(row=tablo_baslik_satir, column=col_toplam_gun, value="Toplam\nGün")
        cell_tg.font = font_baslik
        cell_tg.fill = fill_baslik
        cell_tg.alignment = align_center
        cell_tg.border = thin_border
        ws.column_dimensions[cell_tg.column_letter].width = 10

        cell_tm = ws.cell(row=tablo_baslik_satir, column=col_toplam_mesai, value="Toplam\nMesai")
        cell_tm.font = font_baslik
        cell_tm.fill = fill_baslik
        cell_tm.alignment = align_center
        cell_tm.border = thin_border
        ws.column_dimensions[cell_tm.column_letter].width = 10

        current_row = 5

        # 5. PERSONEL DÖNGÜSÜ (HÜCRELER ELLE YAZILMASI İÇİN BOŞ)
        for p in personeller:
            p_gorev = p.get("gorev", "İşçi")
            if "ad_soyad" in p:
                p_ad = p["ad_soyad"]
            else:
                p_ad = f"{p.get('ad', '')} {p.get('soyad', '')}".strip() or "Personel"

            # --- SATIR 1: YOKLAMA SATIRI ---
            ws.row_dimensions[current_row].height = 22 
            
            cell_name1 = ws.cell(row=current_row, column=1, value=p_ad)
            cell_name1.font = font_bold
            cell_name1.alignment = align_left
            cell_name1.border = thin_border

            cell_gorev1 = ws.cell(row=current_row, column=2, value=p_gorev)
            cell_gorev1.font = font_veri
            cell_gorev1.alignment = align_left
            cell_gorev1.border = thin_border

            for g in range(1, gun_sayisi + 1):
                hucre_yoklama = ws.cell(row=current_row, column=g + 2)
                hucre_yoklama.border = thin_border
                hucre_yoklama.value = ""
                
                hafta_gunu = date(secilen_yil, secilen_ay, g).weekday()
                if hafta_gunu == 5: hucre_yoklama.fill = fill_cumartesi
                elif hafta_gunu == 6: hucre_yoklama.fill = fill_pazar

            ws.cell(row=current_row, column=col_toplam_gun).border = thin_border
            ws.cell(row=current_row, column=col_toplam_mesai).border = thin_border

            current_row += 1

            # --- SATIR 2: MESAİ SATIRI ---
            ws.row_dimensions[current_row].height = 20
            
            cell_name2 = ws.cell(row=current_row, column=1, value=" ⏱ Mesai Saati")
            cell_name2.font = Font(name="Segoe UI", size=9, italic=True, color="595959")
            cell_name2.alignment = align_left
            cell_name2.border = thin_border

            cell_gorev2 = ws.cell(row=current_row, column=2, value="")
            cell_gorev2.border = thin_border

            for g in range(1, gun_sayisi + 1):
                hucre_mesai = ws.cell(row=current_row, column=g + 2)
                hucre_mesai.border = thin_border
                hucre_mesai.value = ""
                
                hafta_gunu = date(secilen_yil, secilen_ay, g).weekday()
                if hafta_gunu == 5: hucre_mesai.fill = fill_cumartesi
                elif hafta_gunu == 6: hucre_mesai.fill = fill_pazar

            ws.cell(row=current_row, column=col_toplam_gun).border = thin_border
            ws.cell(row=current_row, column=col_toplam_mesai).border = thin_border

            current_row += 1

        # 6. KURUMSAL İMZA ALANI (En alt sayfaya eklenir)
        current_row += 3
        imza_unvanlar = ["HAZIRLAYAN", "MUHASEBE", "ŞANTİYE ŞEFİ", "PROJE MÜDÜRÜ"]
        sutun_araligi = max(1, gun_sayisi // 4)
        
        for i, unvan in enumerate(imza_unvanlar):
            start_col = 1 + (i * sutun_araligi)
            end_col = start_col + sutun_araligi - 1
            if end_col > col_toplam_mesai:
                end_col = col_toplam_mesai
                
            ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=end_col)
            imza_cell = ws.cell(row=current_row, column=start_col, value=unvan)
            imza_cell.font = font_imza
            imza_cell.alignment = align_center
            
            ws.merge_cells(start_row=current_row + 3, start_column=start_col, end_row=current_row + 3, end_column=end_col)
            cizgi_cell = ws.cell(row=current_row + 3, column=start_col, value="...................................")
            cizgi_cell.font = font_veri
            cizgi_cell.alignment = align_center

        # 7. Kaydet ve Bitir
        wb.save(kayit_yolu)
        messagebox.showinfo("Başarılı", f"A3 Formatında Başlık Tekrarlamalı Müsvette Puantaj Şablonu başarıyla masaüstüne kaydedildi:\n{dosya_adi}")

    except Exception as e:
        messagebox.showerror("Hata", f"Müsvette şablonu oluşturulurken bir hata meydana geldi:\n{str(e)}")