import customtkinter as ctk
import calendar
import os
from datetime import date, datetime
import tkinter as tk
from tkinter import messagebox, filedialog

# PDF Üretimi için ReportLab bileşenleri
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# 🟢 Excel Üretimi için OpenPyXL Bileşenleri
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# 🇹🇷 TÜRKÇE KARAKTER DESTEKLİ FONT KAYDI
try:
    pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
    pdfmetrics.registerFont(TTFont('Arial-Bold', 'arialbd.ttf'))
    PDF_FONT = "Arial"
    PDF_FONT_BOLD = "Arial-Bold"
except Exception:
    PDF_FONT = "Helvetica"
    PDF_FONT_BOLD = "Helvetica-Bold"

class MaasArayuz(ctk.CTkFrame):

    def __init__(self, master, personel_getir_func=None, veri_getir_func=None, personel_getir=None, yoklama_getir=None, **kwargs):
        super().__init__(master, fg_color="#f5f6fa", **kwargs)

        self.personel_getir = personel_getir_func or personel_getir
        self.yoklama_getir = veri_getir_func or yoklama_getir

        self.personeller = []
        self.secili_personel_idleri = set()
        self.tumunu_sec_durum = False
        self.checkbox_widgets = {}
        self.odendi_durumlari = {}
        self.kesinti_degerleri = {}
        self.kart_widgetlari = {}

        self.resmi_tatiller = {
            "2026-01-01", "2026-04-23", "2026-05-01", "2026-05-19",
            "2026-07-15", "2026-08-30", "2026-10-29"
        }

        self.build_ui()

    def _tarih_parse(self, value):
        if not value or str(value).lower() in ("none", "null", "-"):
            return None
        try:
            return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
        except Exception:
            return None

    def _personel_tarihleri(self, personel):
        giris = self._tarih_parse(
            personel.get("ise_giris_tarihi") or personel.get("ise_giris") or personel.get("sigorta_giris_tarihi")
        )
        cikis = self._tarih_parse(personel.get("isten_cikis_tarihi") or personel.get("cikis_tarihi"))
        return giris, cikis

    def _gun_aktif_mi(self, gun_tarihi, giris, cikis):
        if giris and gun_tarihi < giris:
            return False
        if cikis and gun_tarihi > cikis:
            return False
        return True

    def _hakedis_gunu(self, hak_gun, aktif_gun, eksik_gun):
        if aktif_gun > 0 and hak_gun == aktif_gun and eksik_gun == 0 and aktif_gun >= 28:
            return 30
        return hak_gun

    def build_ui(self):
        # Üst Filtreleme ve Kontrol Paneli
        ust = ctk.CTkFrame(self, fg_color="white", height=60, corner_radius=8)
        ust.pack(fill="x", padx=15, pady=(15, 5))
        ust.pack_propagate(False)

        lbl_ay = ctk.CTkLabel(ust, text="Ay:", font=("Segoe UI", 12, "bold"))
        lbl_ay.pack(side="left", padx=(15, 5), pady=15)
         
        self.ay = ctk.CTkComboBox(ust, values=[str(i) for i in range(1, 13)], width=70)
        self.ay.set(str(date.today().month))
        self.ay.pack(side="left", pady=15)

        lbl_yil = ctk.CTkLabel(ust, text="Yıl:", font=("Segoe UI", 12, "bold"))
        lbl_yil.pack(side="left", padx=(15, 5), pady=15)

        self.yil = ctk.CTkComboBox(ust, values=["2025", "2026", "2027"], width=90)
        self.yil.set(str(date.today().year))
        self.yil.pack(side="left", pady=15)

        self.btn_hesapla = ctk.CTkButton(
            ust,
            text="📊 MAAŞLARI HESAPLA",
            font=("Segoe UI", 12, "bold"),
            fg_color="#1F4E78",
            hover_color="#2B5B84",
            command=self.hesapla
        )
        self.btn_hesapla.pack(side="left", padx=15, pady=15)

        lbl_ara = ctk.CTkLabel(ust, text="Personel Ara:", font=("Segoe UI", 12, "bold"), text_color="#475569")
        lbl_ara.pack(side="left", padx=(15, 5), pady=15)

        self.arama_entry = ctk.CTkEntry(
            ust, 
            placeholder_text="İsim veya soyisim yazın...", 
            width=150,
            font=("Segoe UI", 12)
        )
        self.arama_entry.pack(side="left", pady=15)
        self.arama_entry.bind("<KeyRelease>", lambda event: self.listeyi_filtrele())

        # Sağ Taraf Buton Grubu
        self.btn_tumunu_sec = ctk.CTkButton(
            ust,
            text="☑️ TÜMÜNÜ SEÇ",
            font=("Segoe UI", 11, "bold"),
            fg_color="#64748B",
            hover_color="#475569",
            width=110,
            command=self.tumunu_sec_tetikle
        )
        self.btn_tumunu_sec.pack(side="right", padx=(5, 15), pady=15)

        # 🟢 EXCEL RAPORU OLUŞTUR Butonu
        self.btn_excel_rapor = ctk.CTkButton(
            ust,
            text="🟢 EXCEL RAPORU OLUŞTUR",
            font=("Segoe UI", 11, "bold"),
            fg_color="#15803D",
            hover_color="#166534",
            width=160,
            command=self.excel_rapor_uret
        )
        self.btn_excel_rapor.pack(side="right", padx=5, pady=15)

        # 📄 PDF Olarak Raporla Butonu
        self.btn_pdf_rapor = ctk.CTkButton(
            ust,
            text="📄 PDF RAPORLA",
            font=("Segoe UI", 11, "bold"),
            fg_color="#991B1B",
            hover_color="#7F1D1D",
            width=120,
            command=self.pdf_rapor_uret
        )
        self.btn_pdf_rapor.pack(side="right", padx=5, pady=15)

        # Maliyet Özet Paneli (Dashboard)
        self.dashboard_frame = ctk.CTkFrame(self, fg_color="transparent", height=50)
        self.dashboard_frame.pack(fill="x", padx=15, pady=5)
        
        self.lbl_dash_maas = ctk.CTkLabel(self.dashboard_frame, text="Toplam Hakediş: 0.00 TL", font=("Segoe UI", 11, "bold"), fg_color="#E2E8F0", text_color="#1E293B", corner_radius=6, height=35, width=190)
        self.lbl_dash_maas.pack(side="left", padx=(0, 10))

        self.lbl_dash_mesai = ctk.CTkLabel(self.dashboard_frame, text="Toplam Mesai: 0.00 TL", font=("Segoe UI", 11, "bold"), fg_color="#DBEAFE", text_color="#1E40AF", corner_radius=6, height=35, width=190)
        self.lbl_dash_mesai.pack(side="left", padx=5)

        self.lbl_dash_kesinti = ctk.CTkLabel(self.dashboard_frame, text="Toplam Kesinti: 0.00 TL", font=("Segoe UI", 11, "bold"), fg_color="#FEE2E2", text_color="#991B1B", corner_radius=6, height=35, width=190)
        self.lbl_dash_kesinti.pack(side="left", padx=5)

        self.lbl_dash_net = ctk.CTkLabel(self.dashboard_frame, text="Net Ödenecek (Banka): 0.00 TL", font=("Segoe UI", 12, "bold"), fg_color="#D1FAE5", text_color="#065F46", corner_radius=6, height=35, width=220)
        self.lbl_dash_net.pack(side="left", padx=5)

        # Personel Kartlarının Listeleneceği Scrollable Frame
        self.area = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.area.pack(fill="both", expand=True, padx=15, pady=(0, 15))

    def hesapla(self):
        self.secili_personel_idleri.clear()
        self.checkbox_widgets.clear()
        self.kart_widgetlari.clear()
        self.tumunu_sec_durum = False
        self.btn_tumunu_sec.configure(text="☑️ TÜMÜNÜ SEÇ", fg_color="#64748B")

        if self.personel_getir:
            self.personeller = self.personel_getir()
        else:
            return

        self.listeyi_filtrele()

    def listeyi_filtrele(self):
        for w in self.area.winfo_children():
            w.destroy()

        ay = int(self.ay.get())
        yil = int(self.yil.get())
        gun_sayisi = calendar.monthrange(yil, ay)[1]
        arama_terimi = self.arama_entry.get().strip().lower()

        self.toplam_hakedis_maas = 0.0
        self.toplam_hakedis_mesai = 0.0
        self.toplam_kesinti = 0.0
        self.toplam_net_odenecek = 0.0

        for p in self.personeller:
            p_id = p["id"]
            tam_isim = f"{p['ad']} {p['soyad']}".lower()
            
            if arama_terimi and arama_terimi not in tam_isim:
                continue

            ise_giris_tarihi, isten_cikis_tarihi = self._personel_tarihleri(p)

            veri = self.yoklama_getir(p_id, yil, ay) if self.yoklama_getir else []

            gun_map = {}
            mesai_map = {}
            for v in veri:
                tarih_parca = str(v["tarih"]).split("-")
                if len(tarih_parca) == 3:
                    g = int(tarih_parca[2])
                    gun_map[g] = v["durum"]
                    mesai_map[g] = float(v.get("mesai_saati") or 0)

            sayac_geldi = 0
            sayac_yok = 0
            sayac_izin = 0
            sayac_rapor = 0
            sayac_pazar = 0
            sayac_tatil = 0
            sayac_girilmedi = 0
            aktif_gun = 0
            toplam_mesai = 0

            is_odendi = self.odendi_durumlari.get(p_id, False)
            bg_color_card = "#F8FAFC" if is_odendi else "white"
            border_color_card = "#CBD5E1" if is_odendi else "#E2E8F0"

            row = ctk.CTkFrame(self.area, fg_color=bg_color_card, corner_radius=8, border_width=1, border_color=border_color_card, height=175)
            row.pack(fill="x", pady=5, padx=5)
            row.pack_propagate(False)
            self.kart_widgetlari[p_id] = row

            # Seçim Paneli
            secim_frame = ctk.CTkFrame(row, fg_color="transparent", width=40)
            secim_frame.pack(side="left", padx=(10, 0), fill="y")
            secim_frame.pack_propagate(False)

            cb = ctk.CTkCheckBox(
                secim_frame, 
                text="", 
                width=20, 
                checkbox_width=20, 
                checkbox_height=20,
                command=lambda pid=p_id: self.checkbox_tetiklendi(pid)
            )
            cb.pack(expand=True)
            self.checkbox_widgets[p_id] = cb
            if p_id in self.secili_personel_idleri:
                cb.select()

            # Sol Panel: Personel Künyesi
            sol = ctk.CTkFrame(row, fg_color="transparent")
            sol.pack(side="left", padx=10, pady=12, fill="y")

            lbl_isim = ctk.CTkLabel(sol, text=f"👷 {p['ad']} {p['soyad']}", width=180, anchor="w", font=("Segoe UI", 13, "bold"), text_color="#1E293B" if not is_odendi else "#64748B")
            lbl_isim.pack(anchor="w", pady=(2, 2))

            tc_no = p.get('tc_no', '-')
            gorev = p.get('gorev', p.get('unvan', '-'))
            base_maas = float(p.get('maas') or 0)

            ctk.CTkLabel(sol, text=f"TC: {tc_no}  | Gör: {gorev}", font=("Segoe UI", 11), text_color="#64748B").pack(anchor="w")
            ctk.CTkLabel(sol, text=f"Maaş: {base_maas:,.2f} TL", font=("Segoe UI", 11, "bold"), text_color="#334155").pack(anchor="w", pady=(2, 0))

            # Cetvel Alanı
            cetvel_ana_frame = ctk.CTkFrame(row, fg_color="transparent")
            cetvel_ana_frame.pack(side="left", padx=10, fill="both", expand=True)

            tarih_row = ctk.CTkFrame(cetvel_ana_frame, fg_color="transparent")
            tarih_row.pack(fill="x", pady=(12, 1))

            yoklama_row = ctk.CTkFrame(cetvel_ana_frame, fg_color="transparent")
            yoklama_row.pack(fill="x", pady=1)

            mesai_row = ctk.CTkFrame(cetvel_ana_frame, fg_color="transparent")
            mesai_row.pack(fill="x", pady=1)

            finans_satir_row = ctk.CTkFrame(cetvel_ana_frame, fg_color="transparent")
            finans_satir_row.pack(fill="x", pady=(3, 10))

            for g in range(1, gun_sayisi + 1):
                guncel_tarih = date(yil, ay, g)
                tarih_kontrol_str = f"{yil}-{ay:02d}-{g:02d}"
                weekday = guncel_tarih.weekday()
                
                ctk.CTkLabel(tarih_row, text=f"{g:02d}", width=21, height=16, font=("Segoe UI", 8, "bold"), text_color="#64748B" if weekday != 6 else "#EA580C").pack(side="left", padx=1)

                durum = gun_map.get(g)
                guncel_mesai = float(mesai_map.get(g) or 0)
                toplam_mesai += guncel_mesai

                henuz_ise_girmedi = not self._gun_aktif_mi(guncel_tarih, ise_giris_tarihi, isten_cikis_tarihi)
                if not henuz_ise_girmedi:
                    aktif_gun += 1

                if henuz_ise_girmedi:
                    txt, color = "-", "#94A3B8"
                elif durum == "Geldi":
                    txt, color = "G", "#16A34A"
                    sayac_geldi += 1
                elif durum == "İzinli":
                    txt, color = "İ", "#EAB308"
                    sayac_izin += 1
                elif durum == "Raporlu":
                    txt, color = "R", "#EF4444"
                    sayac_rapor += 1
                elif durum == "Yok":
                    txt, color = "Y", "#DC2626"
                    sayac_yok += 1
                elif weekday == 6:
                    txt, color = "P", "#EA580C"
                    sayac_girilmedi += 1
                elif tarih_kontrol_str in self.resmi_tatiller:
                    txt, color = "T", "#8B5CF6"
                    sayac_girilmedi += 1
                else:
                    txt, color = "?", "#94A3B8"
                    sayac_girilmedi += 1

                final_color = "#CBD5E1" if is_odendi else color

                ctk.CTkLabel(yoklama_row, text=txt, width=21, height=21, font=("Segoe UI", 9, "bold"), fg_color=final_color, text_color="white", corner_radius=4).pack(side="left", padx=1)

                mesai_txt = str(int(guncel_mesai)) if guncel_mesai.is_integer() else str(guncel_mesai)
                mesai_bg = "#3B82F6" if guncel_mesai > 0 else "#F1F5F9"
                mesai_fg = "white" if guncel_mesai > 0 else "#94A3B8"
                final_mesai_bg = "#E2E8F0" if is_odendi else mesai_bg

                ctk.CTkLabel(mesai_row, text=mesai_txt if guncel_mesai > 0 else "-", width=21, height=21, font=("Segoe UI", 9, "bold"), fg_color=final_mesai_bg, text_color=mesai_fg, corner_radius=4).pack(side="left", padx=1)

            hak_edilen_toplam_gun = self._hakedis_gunu(
                sayac_geldi + sayac_pazar + sayac_tatil + sayac_izin + sayac_rapor,
                aktif_gun,
                sayac_yok + sayac_girilmedi,
            )

            ozet_panel = ctk.CTkFrame(row, fg_color="#F8FAFC", border_width=1, border_color="#E2E8F0", width=140)
            ozet_panel.pack(side="left", padx=10, pady=8, fill="y")
            ozet_panel.pack_propagate(False)

            ctk.CTkLabel(ozet_panel, text=f"Top. Gün: {hak_edilen_toplam_gun}", font=("Segoe UI", 10, "bold"), text_color="#1E293B", anchor="w").pack(fill="x", padx=8, pady=(4,1))
            ctk.CTkLabel(ozet_panel, text=f"Top. Mesai: {int(toplam_mesai)} sa", font=("Segoe UI", 10), text_color="#2563EB", anchor="w").pack(fill="x", padx=8, pady=1)
            ctk.CTkLabel(ozet_panel, text=f"Yok Gün: {sayac_yok}", font=("Segoe UI", 10), text_color="#DC2626", anchor="w").pack(fill="x", padx=8, pady=1)
            ctk.CTkLabel(ozet_panel, text=f"İzin:{sayac_izin} | Rap:{sayac_rapor}", font=("Segoe UI", 9), text_color="#7C3AED", anchor="w").pack(fill="x", padx=8, pady=(1,4))

            # Finansal Hakediş Hesaplamaları
            gunluk_ucret = base_maas / 30
            saatlik_ucret = gunluk_ucret / 7.5

            hakedis_maas = max(0.0, hak_edilen_toplam_gun * gunluk_ucret)
            hakedis_mesai = toplam_mesai * saatlik_ucret * 1.5
            
            aktif_kesinti = self.kesinti_degerleri.get(p_id, 0.0)
            net_toplam_maas = (hakedis_maas + hakedis_mesai) - aktif_kesinti

            self.toplam_hakedis_maas += hakedis_maas
            self.toplam_hakedis_mesai += hakedis_mesai
            self.toplam_kesinti += aktif_kesinti
            self.toplam_net_odenecek += net_toplam_maas

            raw_iban = p.get('iban_no', '').strip()
            iban_gosterim = raw_iban if (raw_iban and raw_iban.lower() != "none") else "IBAN Yok"
            iban_renk = "#475569" if iban_gosterim != "IBAN Yok" else "#EF4444"

            # Alt finans satırı yerleşimi
            ctk.CTkLabel(finans_satir_row, text=f"Maaş: {hakedis_maas:,.2f} TL |", font=("Segoe UI", 11, "bold"), text_color="#475569").pack(side="left", padx=(0, 5))
            ctk.CTkLabel(finans_satir_row, text=f"Mesai: {hakedis_mesai:,.2f} TL |", font=("Segoe UI", 11), text_color="#475569").pack(side="left", padx=5)
            ctk.CTkLabel(finans_satir_row, text="Kesinti/Avans:", font=("Segoe UI", 11), text_color="#991B1B").pack(side="left", padx=(5, 2))
            
            kesinti_entry = ctk.CTkEntry(finans_satir_row, width=65, height=20, font=("Segoe UI", 11))
            kesinti_entry.insert(0, f"{aktif_kesinti:.0f}" if aktif_kesinti > 0 else "0")
            kesinti_entry.pack(side="left", padx=(0, 5))
            
            kesinti_entry.bind("<FocusOut>", lambda event, pid=p_id, ent=kesinti_entry: self.kesinti_guncelle(pid, ent))
            kesinti_entry.bind("<Return>", lambda event, pid=p_id, ent=kesinti_entry: self.kesinti_guncelle(pid, ent))

            ctk.CTkLabel(finans_satir_row, text=f"| IBAN: {iban_gosterim} |", font=("Segoe UI", 11), text_color=iban_renk).pack(side="left", padx=5)
            
            lbl_net_maas = ctk.CTkLabel(finans_satir_row, text=f"NET ALACAK: {net_toplam_maas:,.2f} TL", font=("Segoe UI", 12, "bold"), text_color="#16A34A" if net_toplam_maas > 0 else "#64748B")
            lbl_net_maas.pack(side="left", padx=5)

            # Sağ Panel: Hızlı Aksiyonlar
            sag_bos_alan = ctk.CTkFrame(row, fg_color="transparent", width=140)
            sag_bos_alan.pack(side="right", padx=15, pady=12, fill="y")
            sag_bos_alan.pack_propagate(False)

            ctk.CTkButton(sag_bos_alan, text="📋 IBAN Kopyala", font=("Segoe UI", 10, "bold"), fg_color="#0284C7", hover_color="#0369A1", height=24, command=lambda iban=iban_gosterim: self.iban_kopyala(iban)).pack(fill="x", pady=4)
            
            btn_pay_status = ctk.CTkButton(
                sag_bos_alan,
                text="✖️ Ödeme Kaldır" if is_odendi else "💵 Ödeme Tamam",
                font=("Segoe UI", 10, "bold"),
                fg_color="#475569" if is_odendi else "#16A34A",
                hover_color="#334155" if is_odendi else "#15803D",
                height=24,
                command=lambda pid=p_id: self.odendi_isaretle(pid)
            )
            btn_pay_status.pack(fill="x", pady=4)

        # Dashboard Alanını Güncelleme
        self.lbl_dash_maas.configure(text=f"Toplam Hakediş: {self.toplam_hakedis_maas:,.2f} TL")
        self.lbl_dash_mesai.configure(text=f"Toplam Mesai: {self.toplam_hakedis_mesai:,.2f} TL")
        self.lbl_dash_kesinti.configure(text=f"Toplam Kesinti: {self.toplam_kesinti:,.2f} TL")
        self.lbl_dash_net.configure(text=f"Net Ödenecek (Banka): {self.toplam_net_odenecek:,.2f} TL")

    def kesinti_guncelle(self, personel_id, entry_widget):
        try:
            val = float(entry_widget.get().strip())
            self.kesinti_degerleri[personel_id] = max(0.0, val)
        except ValueError:
            self.kesinti_degerleri[personel_id] = 0.0
        self.listeyi_filtrele()

    def iban_kopyala(self, iban):
        if iban == "IBAN Yok" or not iban:
            return
        self.clipboard_clear()
        self.clipboard_append(iban)
        self.update()
        
        top = ctk.CTkToplevel(self)
        top.geometry("200x50")
        top.attributes("-topmost", True)
        top.overrideredirect(True)
        cx = self.winfo_toplevel().winfo_x() + 300
        cy = self.winfo_toplevel().winfo_y() + 300
        top.geometry(f"200x50+{cx}+{cy}")
        ctk.CTkLabel(top, text="📋 IBAN Kopyalandı!", font=("Segoe UI", 12, "bold"), text_color="#0369A1").pack(expand=True, fill="both")
        self.after(1200, top.destroy)

    def odendi_isaretle(self, personel_id):
        su_an = self.odendi_durumlari.get(personel_id, False)
        self.odendi_durumlari[personel_id] = not su_an
        self.listeyi_filtrele()

    def checkbox_tetiklendi(self, personel_id):
        if personel_id in self.secili_personel_idleri:
            self.secili_personel_idleri.remove(personel_id)
        else:
            self.secili_personel_idleri.add(personel_id)

    def tumunu_sec_tetikle(self):
        if not self.personeller:
            return
        if not self.tumunu_sec_durum:
            for p in self.personeller:
                self.secili_personel_idleri.add(p["id"])
                if p["id"] in self.checkbox_widgets:
                    self.checkbox_widgets[p["id"]].select()
            self.tumunu_sec_durum = True
            self.btn_tumunu_sec.configure(text="✖️ SEÇİMİ KALDIR", fg_color="#EF4444", hover_color="#DC2626")
        else:
            self.secili_personel_idleri.clear()
            for cb in self.checkbox_widgets.values():
                cb.deselect()
            self.tumunu_sec_durum = False
            self.btn_tumunu_sec.configure(text="☑️ TÜMÜNÜ SEÇ", fg_color="#64748B", hover_color="#475569")

    # ================= 📄 PDF BÖLÜMÜ =================
    def pdf_rapor_uret(self):
        if not self.personeller:
            messagebox.showwarning("Uyarı", "Raporlanacak veri bulunamadı! Lütfen önce hesaplama yapın.")
            return

        raporlanacak_personeller = self.get_secili_veya_tumu()
        rapor_turu_metni = f"Seçili {len(raporlanacak_personeller)} Personel" if self.secili_personel_idleri else "Tüm Şantiye"

        ay = self.ay.get()
        yil = self.yil.get()
        varsayilan_ad = f"Santiye_Maas_Bordrosu_{yil}_{ay}.pdf"
        
        dosya_yolu = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF Dosyası", "*.pdf")],
            initialfile=varsayilan_ad,
            title="Maaş Bordrosunu Kaydet"
        )
        
        if not dosya_yolu:
            return

        try:
            doc = SimpleDocTemplate(
                dosya_yolu,
                pagesize=landscape(A4),
                rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20
            )
            
            styles = getSampleStyleSheet()
            
            style_title = ParagraphStyle('TitleStyle', fontName=PDF_FONT_BOLD, fontSize=15, leading=18, textColor=colors.HexColor('#1F4E78'))
            style_meta = ParagraphStyle('MetaStyle', fontName=PDF_FONT, fontSize=9, leading=13, textColor=colors.HexColor('#475569'), alignment=2)
            style_cell = ParagraphStyle('CellStyle', fontName=PDF_FONT, fontSize=8.5, leading=11, textColor=colors.HexColor('#1E293B'), alignment=1)
            style_cell_left = ParagraphStyle('CellLeftStyle', fontName=PDF_FONT, fontSize=8, leading=11, textColor=colors.HexColor('#1E293B'))
            style_cell_bold = ParagraphStyle('CellBoldStyle', fontName=PDF_FONT_BOLD, fontSize=9, leading=12, textColor=colors.HexColor('#1E293B'))
            style_th = ParagraphStyle('ThStyle', fontName=PDF_FONT_BOLD, fontSize=8.5, leading=11, textColor=colors.white, alignment=1)
            
            sub_style_t = ParagraphStyle('SubStyleT', fontName=PDF_FONT_BOLD, fontSize=6.5, leading=8, alignment=1, textColor=colors.HexColor('#475569'))
            sub_style_t_sun = ParagraphStyle('SubStyleTSun', fontName=PDF_FONT_BOLD, fontSize=6.5, leading=8, alignment=1, textColor=colors.HexColor('#EA580C'))
            sub_style_lbl = ParagraphStyle('SubStyleLbl', fontName=PDF_FONT_BOLD, fontSize=6.5, leading=8, alignment=0, textColor=colors.HexColor('#1E293B'))

            hikaye = []

            pdf_toplam_hakedis_maas = 0.0
            pdf_toplam_hakedis_mesai = 0.0
            pdf_toplam_kesinti = 0.0
            pdf_toplam_net_odenecek = 0.0

            gun_sayisi_ay = calendar.monthrange(int(yil), int(ay))[1]
            
            tablo_icerik = [[
                Paragraph("No", style_th),
                Paragraph("Personel Künyesi", style_th),
                Paragraph("Hak<br/>Gün", style_th),
                Paragraph("Hakediş", style_th),
                Paragraph("Mesai<br/>(Sa)", style_th),
                Paragraph("Mesai<br/>Kazanç", style_th),
                Paragraph("Kesinti", style_th),
                Paragraph("Sembolik Yoklama & Mesai Çizelge Barı (Hizalı Cetvel)", style_th),
                Paragraph("NET ALACAK", style_th)
            ]]

            sira = 1
            for p in raporlanacak_personeller:
                p_id = p["id"]
                veri = self.yoklama_getir(p_id, int(yil), int(ay)) if self.yoklama_getir else []
                
                gun_map = {}
                mesai_map = {}
                for v in veri:
                    tarih_parca = str(v["tarih"]).split("-")
                    if len(tarih_parca) == 3:
                        g = int(tarih_parca[2])
                        gun_map[g] = v["durum"]
                        mesai_map[g] = float(v.get("mesai_saati") or 0)

                sayac_geldi, sayac_yok, sayac_izin, sayac_rapor, sayac_pazar, sayac_tatil, toplam_mesai = 0, 0, 0, 0, 0, 0, 0
                
                sub_row_t = [Paragraph("<b>T:</b>", sub_style_lbl)]
                sub_row_y = [Paragraph("<b>Y:</b>", sub_style_lbl)]
                sub_row_m = [Paragraph("<b>M:</b>", sub_style_lbl)]

                for g in range(1, gun_sayisi_ay + 1):
                    guncel_tarih = date(int(yil), int(ay), g)
                    tarih_kontrol_str = f"{yil}-{int(ay):02d}-{g:02d}"
                    weekday = guncel_tarih.weekday()
                    
                    durum = gun_map.get(g)
                    guncel_mesai = float(mesai_map.get(g) or 0)
                    toplam_mesai += guncel_mesai

                    if weekday == 6:
                        sub_row_t.append(Paragraph(f"{g:02d}", sub_style_t_sun))
                    else:
                        sub_row_t.append(Paragraph(f"{g:02d}", sub_style_t))

                    if durum == "Geldi":
                        y_char = '<font color="#16A34A"><b>G</b></font>'; sayac_geldi += 1
                    elif durum == "İzinli":
                        y_char = '<font color="#EAB308"><b>İ</b></font>'; sayac_izin += 1
                    elif durum == "Raporlu":
                        y_char = '<font color="#EF4444"><b>R</b></font>'; sayac_rapor += 1
                    elif durum == "Yok":
                        y_char = '<font color="#DC2626"><b>Y</b></font>'; sayac_yok += 1
                    elif weekday == 6:
                        y_char = '<font color="#EA580C"><b>P</b></font>'
                    elif tarih_kontrol_str in self.resmi_tatiller:
                        y_char = '<font color="#8B5CF6"><b>T</b></font>'
                    else:
                        y_char = '<font color="#94A3B8">-</font>'
                    sub_row_y.append(Paragraph(y_char, sub_style_t))

                    m_char = f'<font color="#3B82F6"><b>{int(guncel_mesai)}</b></font>' if guncel_mesai > 0 else '<font color="#CBD5E1">-</font>'
                    sub_row_m.append(Paragraph(m_char, sub_style_t))

                for g in range(gun_sayisi_ay + 1, 32):
                    sub_row_t.append(Paragraph("", sub_style_t))
                    sub_row_y.append(Paragraph("", sub_style_t))
                    sub_row_m.append(Paragraph("", sub_style_t))

                sub_widths = [15] + [9] * 31
                sub_table = Table([sub_row_t, sub_row_y, sub_row_m], colWidths=sub_widths)
                sub_table.setStyle(TableStyle([
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('LEFTPADDING', (0,0), (-1,-1), 0),
                    ('RIGHTPADDING', (0,0), (-1,-1), 0),
                    ('TOPPADDING', (0,0), (-1,-1), 1),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 1),
                ]))

                hak_edilen_toplam_gun = sayac_geldi + sayac_pazar + sayac_tatil + sayac_izin + sayac_rapor
                if hak_edilen_toplam_gun >= (gun_sayisi_ay - sayac_yok) and sayac_yok == 0:
                    hak_edilen_toplam_gun = 30

                base_maas = float(p.get('maas') or 0)
                gunluk_ucret = base_maas / 30
                saatlik_ucret = gunluk_ucret / 7.5

                hakedis_maas = max(0.0, hak_edilen_toplam_gun * gunluk_ucret)
                hakedis_mesai = toplam_mesai * saatlik_ucret * 1.5
                aktif_kesinti = self.kesinti_degerleri.get(p_id, 0.0)
                net_toplam_maas = (hakedis_maas + hakedis_mesai) - aktif_kesinti

                pdf_toplam_hakedis_maas += hakedis_maas
                pdf_toplam_hakedis_mesai += hakedis_mesai
                pdf_toplam_kesinti += aktif_kesinti
                pdf_toplam_net_odenecek += net_toplam_maas

                raw_iban = p.get('iban_no', '').strip()
                iban_gosterim = raw_iban if (raw_iban and raw_iban.lower() != "none") else "-"

                kunye_html = f"<b>{p['ad']} {p['soyad']}</b><br/><font size=7 color='#475569'>Gör: {p.get('unvan', p.get('gorev','-'))}<br/>IBAN: {iban_gosterim}<br/>Maaş: {base_maas:,.2f} TL</font>"

                tablo_icerik.append([
                    Paragraph(str(sira), style_cell),
                    Paragraph(kunye_html, style_cell_left),
                    Paragraph(str(hak_edilen_toplam_gun), style_cell),
                    Paragraph(f"{hakedis_maas:,.2f}", style_cell),
                    Paragraph(str(int(toplam_mesai)), style_cell),
                    Paragraph(f"{hakedis_mesai:,.2f}", style_cell),
                    Paragraph(f"{aktif_kesinti:,.2f}", style_cell),
                    sub_table,  
                    Paragraph(f"<b>{net_toplam_maas:,.2f} TL</b>", style_cell)
                ])
                sira += 1

            logo_yolu = os.path.join(r"C:\Users\DELL\source\repos\santiyePROGRAM", "logo-kibritci.png")
            meta_html = f"<b>Dönem:</b> {ay} / {yil}<br/><b>Rapor Türü:</b> {rapor_turu_metni}<br/><b>Rapor Tarihi:</b> {date.today().strftime('%d.%m.%Y')}<br/><b>Çalışan Sayısı:</b> {len(raporlanacak_personeller)}"
            
            if os.path.exists(logo_yolu):
                logo_img = Image(logo_yolu, width=80, height=80)
                baslik_tablo = Table([
                    [logo_img, 
                     Paragraph("<b>KİBRİTÇİ İNŞAAT ERP SİSTEMİ</b><br/>ŞANTİYE MAAŞ BORDRO LİSTESİ", style_title),
                     Paragraph(meta_html, style_meta)]
                ], colWidths=[90, 410, 252])
            else:
                baslik_tablo = Table([
                    ["", 
                     Paragraph("<b>KİBRİTÇİ İNŞAAT ERP SİSTEMİ</b><br/>ŞANTİYE MAAŞ BORDRO LİSTESİ", style_title),
                     Paragraph(meta_html, style_meta)]
                ], colWidths=[10, 490, 252])

            baslik_tablo.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 10)
            ]))
            hikaye.append(baslik_tablo)
            hikaye.append(Spacer(1, 10))

            # Finansal Özet Bandı
            ozet_data = [
                [Paragraph("<b>Toplam Hakediş Maaş</b>", style_cell_bold), 
                 Paragraph("<b>Toplam Hakediş Mesai</b>", style_cell_bold), 
                 Paragraph("<b>Toplam Kesinti / Avans</b>", style_cell_bold), 
                 Paragraph("<b>NET BANKA TALİMATI</b>", style_cell_bold)],
                [f"{pdf_toplam_hakedis_maas:,.2f} TL", f"{pdf_toplam_hakedis_mesai:,.2f} TL", f"{pdf_toplam_kesinti:,.2f} TL", f"{pdf_toplam_net_odenecek:,.2f} TL"]
            ]
            ozet_tablo = Table(ozet_data, colWidths=[188, 188, 188, 188])
            ozet_tablo.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F1F5F9')),
                ('BACKGROUND', (3,0), (3,1), colors.HexColor('#D1FAE5')), 
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
                ('TOPPADDING', (0,0), (-1,-1), 6),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ]))
            hikaye.append(ozet_tablo)
            hikaye.append(Spacer(1, 15))

            # Ana Bordro Tablosu
            ana_tablo = Table(tablo_icerik, colWidths=[20, 125, 30, 55, 30, 55, 50, 310, 77])
            ana_tablo.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E78')), 
                ('ALIGN', (0,0), (-1,0), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]) 
            ]))
            hikaye.append(ana_tablo)
            hikaye.append(Spacer(1, 25))

            # İmza Alanı
            style_imza = ParagraphStyle('ImzaStyle', fontName=PDF_FONT_BOLD, fontSize=9, alignment=1)
            style_imza_alt = ParagraphStyle('ImzaAltStyle', fontName=PDF_FONT, fontSize=9, alignment=1)
            
            imza_veri = [
                [Paragraph("HAZIRLAYAN", style_imza),
                 Paragraph("MUHASEBE", style_imza),
                 Paragraph("ŞANTİYE ŞEFİ", style_imza),
                 Paragraph("PROJE MÜDÜRÜ", style_imza)],
                [Paragraph("<br/><br/>___________________<br/>İmza", style_imza_alt), 
                 Paragraph("<br/><br/>___________________<br/>İmza", style_imza_alt), 
                 Paragraph("<br/><br/>___________________<br/>İmza", style_imza_alt), 
                 Paragraph("<br/><br/>___________________<br/>İmza", style_imza_alt)]
            ]
            imza_tablo = Table(imza_veri, colWidths=[188, 188, 188, 188])
            imza_tablo.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('TOPPADDING', (0,0), (-1,-1), 2),
                ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ]))
            hikaye.append(KeepTogether(imza_tablo))

            doc.build(hikaye)
            messagebox.showinfo("Başarılı", f"Maaş bordro raporu başarıyla oluşturuldu:\n{dosya_yolu}")

        except Exception as e:
            messagebox.showerror("Hata", f"PDF Raporu üretilirken beklenmeyen bir hata oluştu:\n{e}")

    # ================= 🟢 MODÜLER VE KUSURSUZ HİZALANMIŞ EXCEL RAPORLAMA BÖLÜMÜ =================
    def excel_rapor_uret(self):
        if not EXCEL_AVAILABLE:
            messagebox.showerror("Hata", "Excel raporu oluşturmak için 'openpyxl' kütüphanesi gerekli.")
            return

        if not self.personeller:
            messagebox.showwarning("Uyarı", "Raporlanacak veri bulunamadı!")
            return

        raporlanacak_personeller = self.get_secili_veya_tumu()
        ay = self.ay.get()
        yil = self.yil.get()
        varsayilan_ad = f"Santiye_Maas_Bordrosu_{yil}_{ay}.xlsx"

        dosya_yolu = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyası", "*.xlsx")],
            initialfile=varsayilan_ad,
            title="Excel Bordrosunu Kaydet"
        )

        if not dosya_yolu:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Bordro {ay}-{yil}"
            ws.views.sheetView[0].showGridLines = True

            # 🎨 Profesyonel Hücre Stilleri
            font_baslik = Font(name="Arial", size=15, bold=True, color="1F4E78")
            font_sub_baslik = Font(name="Arial", size=10, italic=True, color="475569")
            font_th = Font(name="Arial", size=9, bold=True, color="FFFFFF")
            font_normal = Font(name="Arial", size=9)
            font_bold = Font(name="Arial", size=9, bold=True)
            font_pazar = Font(name="Arial", size=9, bold=True, color="C2410C")
            font_cizelge_lbl = Font(name="Arial", size=8.5, bold=True, color="1E293B")

            fill_th = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            fill_pazar = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")  # Yumuşak Turuncu Şerit
            fill_net = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")    # Yumuşak Yeşil Net Alacak
            fill_lbl_cizelge = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

            border_ince = Border(
                left=Side(style='thin', color='CBD5E1'),
                right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'),
                bottom=Side(style='thin', color='CBD5E1')
            )

            align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            align_left = Alignment(horizontal='left', vertical='center')
            align_right = Alignment(horizontal='right', vertical='center')

            # 🏢 Üst Bilgi Başlıkları
            ws["A1"] = "KİBRİTÇİ İNŞAAT ERP SİSTEMİ"
            ws["A1"].font = font_baslik
            ws["A2"] = f"ŞANTİYE MAAŞ BORDRO LİSTESİ - DÖNEM: {ay}/{yil} | Rapor Tarihi: {date.today().strftime('%d.%m.%Y')}"
            ws["A2"].font = font_sub_baslik

            # 📊 Sütun Düzeni Yapılandırması
            headers = ["No", "Ad Soyad", "Görevi", "TC Kimlik No", "IBAN No", "Kök Maaş", "Çizelge", "Hak Gün", "Hakediş Maaş", "Mesai (Sa)", "Mesai Kazanç", "Kesinti"]
            gun_sayisi_ay = calendar.monthrange(int(yil), int(ay))[1]

            start_row = 5
            
            # Ana Sütun Başlıklarının Yazılması
            for col_num, h_text in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col_num, value=h_text)
                cell.font = font_th
                cell.fill = fill_th
                cell.alignment = align_center
                cell.border = border_ince

            # Çizelge 1-31 Gün Kolon Başlıkları
            cizelge_start_col = len(headers) + 1
            for g in range(1, gun_sayisi_ay + 1):
                cell = ws.cell(row=start_row, column=cizelge_start_col + g - 1, value=f"{g:02d}")
                cell.font = font_th
                cell.fill = fill_th
                cell.alignment = align_center
                cell.border = border_ince

            # Net Alacak Sütun Başlığı
            net_alacak_col = cizelge_start_col + gun_sayisi_ay
            cell_net_h = ws.cell(row=start_row, column=net_alacak_col, value="NET ALACAK")
            cell_net_h.font = font_th
            cell_net_h.fill = fill_th
            cell_net_h.alignment = align_center
            cell_net_h.border = border_ince

            current_row = start_row + 1
            sira = 1

            # 👷 Personel Satırlarının İnşası
            for p in raporlanacak_personeller:
                p_id = p["id"]
                veri = self.yoklama_getir(p_id, int(yil), int(ay)) if self.yoklama_getir else []

                gun_map = {}
                mesai_map = {}
                for v in veri:
                    tarih_parca = str(v["tarih"]).split("-")
                    if len(tarih_parca) == 3:
                        g = int(tarih_parca[2])
                        gun_map[g] = v["durum"]
                        mesai_map[g] = float(v.get("mesai_saati") or 0)

                sayac_geldi, sayac_yok, sayac_izin, sayac_rapor, sayac_pazar, sayac_tatil, toplam_mesai = 0, 0, 0, 0, 0, 0, 0
                base_maas = float(p.get('maas') or 0)
                aktif_kesinti = self.kesinti_degerleri.get(p_id, 0.0)

                is_zebra = (sira % 2 == 0)
                row_fill = fill_zebra if is_zebra else PatternFill(fill_type=None)

                # 🔗 1. KISIM: Sol Künye Alanlarının Birleştirilmesi ve Yazılması (3'er Satır Dikey Merge)
                merge_targets = [1, 2, 3, 4, 5, 6]
                for col_idx in merge_targets:
                    ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row + 2, end_column=col_idx)

                ws.cell(row=current_row, column=1, value=sira).alignment = align_center
                ws.cell(row=current_row, column=2, value=f"{p['ad']} {p['soyad']}").alignment = align_left
                ws.cell(row=current_row, column=3, value=p.get('unvan', p.get('gorev','-'))).alignment = align_left
                ws.cell(row=current_row, column=4, value=p.get('tc_no', '-')).alignment = align_center
                ws.cell(row=current_row, column=5, value=p.get('iban_no', '-')).alignment = align_center
                
                cell_km = ws.cell(row=current_row, column=6, value=base_maas)
                cell_km.number_format = '#,##0.00 "TL"'
                cell_km.alignment = align_right

                # 📊 2. KISIM: İndikatör Sütunu (T, Y, M)
                c_lbl_t = ws.cell(row=current_row, column=7, value="T")
                c_lbl_y = ws.cell(row=current_row + 1, column=7, value="Y")
                c_lbl_m = ws.cell(row=current_row + 2, column=7, value="M")
                
                for offset, c_lbl in enumerate([c_lbl_t, c_lbl_y, c_lbl_m]):
                    c_lbl.font = font_cizelge_lbl
                    c_lbl.alignment = align_center
                    c_lbl.fill = fill_lbl_cizelge

                # 🗓️ 3. KISIM: Günlük Döngü ve Koşullu Biçimlendirme
                for g in range(1, gun_sayisi_ay + 1):
                    guncel_tarih = date(int(yil), int(ay), g)
                    tarih_kontrol_str = f"{yil}-{int(ay):02d}-{g:02d}"
                    weekday = guncel_tarih.weekday()
                    
                    durum = gun_map.get(g)
                    guncel_mesai = float(mesai_map.get(g) or 0)
                    toplam_mesai += guncel_mesai

                    cell_t = ws.cell(row=current_row, column=cizelge_start_col + g - 1, value=g)
                    cell_y = ws.cell(row=current_row + 1, column=cizelge_start_col + g - 1)
                    cell_m = ws.cell(row=current_row + 2, column=cizelge_start_col + g - 1, value=int(guncel_mesai) if guncel_mesai > 0 else "-")

                    # Gün Hücrelerinin Ortalanması
                    cell_t.alignment = align_center
                    cell_y.alignment = align_center
                    cell_m.alignment = align_center
                    
                    cell_t.font = font_normal
                    cell_m.font = font_normal

                    # Yoklama Karakter Ataması
                    if durum == "Geldi":
                        cell_y.value = "G"; sayac_geldi += 1; cell_y.font = font_bold
                    elif durum == "İzinli":
                        cell_y.value = "İ"; sayac_izin += 1; cell_y.font = font_bold
                    elif durum == "Raporlu":
                        cell_y.value = "R"; sayac_rapor += 1; cell_y.font = font_bold
                    elif durum == "Yok":
                        cell_y.value = "Y"; sayac_yok += 1; cell_y.font = font_bold
                    elif weekday == 6:
                        cell_y.value = "P"; cell_y.font = font_pazar
                    elif tarih_kontrol_str in self.resmi_tatiller:
                        cell_y.value = "T"; cell_y.font = font_bold
                    else:
                        cell_y.value = "-"; cell_y.font = font_normal

                    # Pazar Günlerini Komple Dikey Şerit Boyama
                    if weekday == 6:
                        cell_t.fill = fill_pazar
                        cell_y.fill = fill_pazar
                        cell_m.fill = fill_pazar

                # 🧮 4. KISIM: Matematiksel Hesaplamalar ve Hak Gün Çözümü
                hak_edilen_toplam_gun = sayac_geldi + sayac_pazar + sayac_tatil + sayac_izin + sayac_rapor
                if hak_edilen_toplam_gun >= (gun_sayisi_ay - sayac_yok) and sayac_yok == 0:
                    hak_edilen_toplam_gun = 30

                gunluk_ucret = base_maas / 30
                saatlik_ucret = gunluk_ucret / 7.5

                hakedis_maas = max(0.0, hak_edilen_toplam_gun * gunluk_ucret)
                hakedis_mesai = toplam_mesai * saatlik_ucret * 1.5
                net_toplam_maas = (hakedis_maas + hakedis_mesai) - aktif_kesinti

                # 💵 5. KISIM: Finansal Hücre Verilerinin Yazılması ve Dikey Birleştirilmesi
                finans_cols = [8, 9, 10, 11, 12]
                for f_col in finans_cols:
                    ws.merge_cells(start_row=current_row, start_column=f_col, end_row=current_row + 2, end_column=f_col)

                ws.cell(row=current_row, column=8, value=hak_edilen_toplam_gun).alignment = align_center
                
                c_hm = ws.cell(row=current_row, column=9, value=hakedis_maas)
                c_hm.number_format = '#,##0.00 "TL"'
                c_hm.alignment = align_right
                
                ws.cell(row=current_row, column=10, value=int(toplam_mesai)).alignment = align_center
                
                c_hk = ws.cell(row=current_row, column=11, value=hakedis_mesai)
                c_hk.number_format = '#,##0.00 "TL"'
                c_hk.alignment = align_right
                
                c_ks = ws.cell(row=current_row, column=12, value=aktif_kesinti)
                c_ks.number_format = '#,##0.00 "TL"'
                c_ks.alignment = align_right

                # 🟢 6. KISIM: NET ALACAK (En Sağ Sütun Birleştirme ve Boyama)
                ws.merge_cells(start_row=current_row, start_column=net_alacak_col, end_row=current_row + 2, end_column=net_alacak_col)
                cell_net = ws.cell(row=current_row, column=net_alacak_col, value=net_toplam_maas)
                cell_net.number_format = '#,##0.00 "TL"'
                cell_net.font = font_bold
                cell_net.fill = fill_net
                cell_net.alignment = align_center

                # 🏁 7. KISIM: Kenarlık Çerçevelerinin ve Zebra Dolguların Hücre Hücre Basılması
                for r_offset in range(3):
                    r_num = current_row + r_offset
                    for c_num in range(1, net_alacak_col + 1):
                        cell_item = ws.cell(row=r_num, column=c_num)
                        cell_item.border = border_ince
                        # Hücrenin kendi özel dolgusu yoksa ve satır zebra ise arka planı boya
                        if cell_item.fill.fill_type is None and row_fill.fill_type is not None:
                            cell_item.fill = row_fill

                current_row += 3
                sira += 1

            # 📐 Sütun Genişliklerini Standartlaştırma ve İçerik Koruma
            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)
                c_idx = col[0].column
                if c_idx >= cizelge_start_col and c_idx < net_alacak_col:
                    ws.column_dimensions[col_letter].width = 4.5  # Gün kolonları ince, simetrik cetvel gibi dursun
                elif c_idx in [2, 5]:  # İsim ve IBAN geniş olsun
                    ws.column_dimensions[col_letter].width = 24
                elif c_idx in [4, 6, 9, 11, 12, net_alacak_col]:  # TC ve Finansal alanlar tam sığsın
                    ws.column_dimensions[col_letter].width = 16
                else:
                    ws.column_dimensions[col_letter].width = 11

            wb.save(dosya_yolu)
            messagebox.showinfo("Başarılı", f"Excel maaş bordro raporu başarıyla oluşturuldu:\n{dosya_yolu}")

        except Exception as e:
            messagebox.showerror("Hata", f"Excel Raporu üretilirken hata oluştu:\n{e}")

    def get_secili_veya_tumu(self):
        if self.secili_personel_idleri:
            return [p for p in self.personeller if p["id"] in self.secili_personel_idleri]
        return self.personeller
